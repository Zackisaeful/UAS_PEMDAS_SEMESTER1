import openpyxl
import os  
from function import print_formatted_separator

FILE_NAME = 'data_spp.xlsx'


# FUNGSI MENAMPILKAN SELURUH LAPORAN SERVIS
def tampil_laporan_servis():
    if not os.path.exists(FILE_NAME):
        print("File data SPP tidak ditemukan.")
        return

    workbook = openpyxl.load_workbook(FILE_NAME)

    # Mengakses sheet Siswa
    if 'Siswa' not in workbook.sheetnames:
        print("Sheet Siswa belum ada.")
        return

    # Ambil data {No Siswa: (Nama, Tahun Angkatan, Kelas)}
    siswa_sheet = workbook['Siswa']
    siswa_data = {str(row[0]): (row[1], row[2], row[3]) for row in siswa_sheet.iter_rows(min_row=2, values_only=True)}

    # Mengakses sheet Pembayaran
    if 'Pembayaran' not in workbook.sheetnames:
        print("Sheet Pembayaran belum ada.")
        return

    pembayaran_sheet = workbook['Pembayaran']
    if pembayaran_sheet.max_row == 1:
        print("Belum ada data pembayaran.")
        return

    print("\nDaftar Laporan Servis Pembayaran SPP:")
    
    for pembayaran_row in pembayaran_sheet.iter_rows(min_row=2, values_only=True):
        kode_tagihan = pembayaran_row[0]
        no_siswa = str(pembayaran_row[1])  
        tanggal_pembayaran = pembayaran_row[2]
        jumlah_pembayaran = pembayaran_row[3]
        uang_diterima = pembayaran_row[4]
        kembalian = pembayaran_row[5]

        # Debugging
        # print(f"Nomor Siswa: {no_siswa}")

        # Mencari siswa berdasarkan nomor siswa
        if no_siswa in siswa_data:
            nama_siswa, tahun_angkatan, kelas = siswa_data[no_siswa]
        else:
            nama_siswa, tahun_angkatan, kelas = "Data siswa tidak ditemukan", "", "", "", ""

        # Menampilkan informasi
        print("\nKode Pembayaran:", kode_tagihan)
        print("Nomor Siswa:", no_siswa)
        print("Nama Siswa:", nama_siswa)
        print("Tahun Angkatan:", tahun_angkatan)
        print("Kelas:", kelas)
        print("Tanggal Pembayaran:", tanggal_pembayaran)  
        print("Jumlah Pembayaran:", jumlah_pembayaran)
        print("Uang Diterima:", uang_diterima)
        print("Kembalian:", kembalian)
        print_formatted_separator()

    workbook.close()



# FUNGSI MENAMPILKAN LAPORAN SERVIS PER SISWA
# Menampilkan laporan servis per siswa.
def tampil_laporan_servis_siswa():
    nomor_siswa = input("Masukkan nomor siswa untuk menampilkan laporan: ")

    if not os.path.exists(FILE_NAME):
        print("File data SPP tidak ditemukan.")
        return

    workbook = openpyxl.load_workbook(FILE_NAME)

    # Mengakses sheet Siswa
    if 'Siswa' not in workbook.sheetnames:
        print("Sheet Siswa belum ada.")
        return

    # Ambil data (Nama, Tahun Angkatan, Kelas)}
    siswa_sheet = workbook['Siswa']
    siswa_data = {str(row[0]): (row[1], row[2], row[3]) for row in siswa_sheet.iter_rows(min_row=2, values_only=True)} 

    # Mengakses sheet Pembayaran
    if 'Pembayaran' not in workbook.sheetnames:
        print("Sheet Pembayaran belum ada.")
        return

    pembayaran_sheet = workbook['Pembayaran']
    if pembayaran_sheet.max_row == 1:
        print("Belum ada data pembayaran.")
        return

    print(f"\nLaporan Servis Pembayaran SPP untuk Siswa Nomor: {nomor_siswa}")
    
    siswa_found = False
    for row in pembayaran_sheet.iter_rows(min_row=2, values_only=True):
        kode_tagihan = row[0]
        no_siswa = str(row[1])  
        tanggal_pembayaran = row[2]
        jumlah_pembayaran = row[3]
        uang_diterima = row[4]
        kembalian = row[5]

        # Cek apakah nomor siswa cocok
        if no_siswa == nomor_siswa:
            siswa_found = True
            # Ambil informasi siswa
            if no_siswa in siswa_data:
                nama_siswa, tahun_angkatan, kelas = siswa_data[no_siswa]
            else:
                nama_siswa, tahun_angkatan, kelas = "Data siswa tidak ditemukan", "", "", "", ""

            # Menampilkan informasi
            print("\nKode Pembayaran:", kode_tagihan)
            print("Nomor Siswa:", no_siswa)
            print("Nama Siswa:", nama_siswa)
            print("Tahun Angkatan:", tahun_angkatan)
            print("Kelas:", kelas)
            print("Tanggal Pembayaran:", tanggal_pembayaran)  
            print("Jumlah Pembayaran:", jumlah_pembayaran)
            print("Uang Diterima:", uang_diterima)
            print("Kembalian:", kembalian)
            print_formatted_separator()

    if not siswa_found:
        print("Tidak ada laporan untuk siswa dengan nomor tersebut.")
    
    workbook.close()
    


# Menu untuk menampilkan laporan
def menu_laporan_servis():
    while True:
        print("\nMenu Tagihan:")
        print("1. Tampilkan seluruh laporan servis pembayaran SPP")
        print("2. Tampilkan laporan servis per siswa")
        print("3. Kembali ke menu utama")
        sub_pilihan = input("Pilih menu: ")

        if sub_pilihan == '1':
            tampil_laporan_servis()
        elif sub_pilihan == '2':
            tampil_laporan_servis_siswa()
        elif sub_pilihan == '3':
            break
        else:
            print("Pilihan tidak valid. Silakan coba lagi.")
