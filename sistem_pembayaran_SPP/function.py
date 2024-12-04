import openpyxl
import os
import random
import string
from datetime import datetime 

# NAMA FILE TEMPAT MENYIMPAN DATA
FILE_NAME = 'data_spp.xlsx'

# FUNGSI MENANGANI SHEET
# Menangani pembuatan sheet baru jika sheet tidak tersedia
def create_sheet_if_not_exists(workbook, sheet_name, header=None):
    if sheet_name not in workbook.sheetnames:
        sheet = workbook.create_sheet(sheet_name)
        if header:
            sheet.append(header)
    return workbook[sheet_name]


# FUNGSI MENENTUKAN TAHUN ANGKATAN
# Mengambil input tahun angkatan siswa dari pengguna.
def tahun_angkatan():
    while True:
        print("\nTahun angkatan siswa:")
        print("1. 2021")
        print("2. 2022")
        print("3. 2023")
        print("4. 2024")
        nomor_angkatan = input("Pilih Tahun angkatan sesuai nomor: ")
        print("---------------------------------------")

        if nomor_angkatan == '1':
            return "2021"
        elif nomor_angkatan == '2':
            return "2022"
        elif nomor_angkatan == '3':
            return "2023"
        elif nomor_angkatan == '4':
            return "2024"
        else:
            print("Input tidak valid")     



# FUNGSI GENERATE NOMOR SISWA 
# Menghasilkan nomor siswa baru secara otomatis.
# Mencari nomor siswa tertinggi yang ada di sheet 'Siswa' dan mengembalikan nomor berikutnya untuk siswa baru.
def generate_no_siswa(workbook):
    if not os.path.exists(FILE_NAME):
        return 1  # Jika file tidak ada, mulai dari 1

    sheet = workbook['Siswa']
    max_no_siswa = 0

    for row in sheet.iter_rows(min_row=2, values_only=True):
        if row[0] is not None:
            max_no_siswa = max(max_no_siswa, int(row[0]))

    return max_no_siswa + 1 


# FUNGSI GENERATE KODE TAGIHAN
# Membuat kode tagihan yang unik
def generate_kode_tagihan(workbook):
    existing_codes = set()

    # Mengambil semua kode tagihan yang sudah ada
    if os.path.exists(FILE_NAME):
        workbook = openpyxl.load_workbook(FILE_NAME)
        if 'Tagihan' in workbook.sheetnames:
            sheet = workbook['Tagihan']
            for row in sheet.iter_rows(min_row=2, values_only=True):
                if row[0] is not None:
                    existing_codes.add(row[0])  # Menyimpan kode yang sudah ada

    while True:
        # Menghasilkan kode tagihan acak dengan kombinasi huruf dan angka
        kode_tagihan = ''.join(random.choices(string.ascii_uppercase + string.digits, k=4))
        if kode_tagihan not in existing_codes:
            return kode_tagihan  
        

# FUNGSI MENANGANI BULAN SPP
# Menangani input bulan spp apakah ingin diisi secara otomatis atau tidak.
# Jika diisi secara otomatis maka bulan diisi sesuai tagihan dibuat
def pilih_bulan():
    bulan_list = [
        "Januari", "Februari", "Maret", "April",
        "Mei", "Juni", "Juli", "Agustus",
        "September", "Oktober", "November", "Desember"
    ]

    pesan = "\nApakah bulan ingin diisi secara otomatis? [y/n]. jika iya bulan akan di isi sesuai waktu pembuatan penagihan: "    

    metode_pilihan = input(pesan)

    if metode_pilihan.lower() == "y":
        nomor_bulan = int(datetime.today().strftime('%m')) 
        bulan_list = bulan_list[nomor_bulan - 1]  

    elif metode_pilihan.lower() == "n":
        while True:
            print("\nPilih Bulan:")
            for i, bulan in enumerate(bulan_list, start=1):
                print(f"{i}. {bulan}")
            
            pilihan = input("Pilih bulan sesuai nomor: ")
            
            if pilihan.isdigit() and 1 <= int(pilihan) <= 12:
                bulan_list = bulan_list[int(pilihan) - 1]  
                break
            else:
                print("\nInput tidak valid. Silakan coba lagi.")
    else:
        print("\nInput tidak valid. Silakan masukkan 'y' atau 'n'.")
        return pilih_bulan()
    print_formatted_separator(len(pesan))
    return bulan_list



# FUNGSI MENENTUKAN JUMLAH TAGIHAN 
# User menentukan jumlah tagihan diisi otomatis atau manual
# Jika otomatis maka jumlah tagihan diisi sesuai anggkatan siswa
def tentukan_tagihan(angkatan_siswa):
    pesan = "\nApakah jumlah tagihan ingin diisi secara otomatis? [y/n]. jika iya tagihan akan di isi sesuai angkatan siswa: "
    metode_pilihan = input(pesan)

    jumlah_tagihan = None

    if metode_pilihan.lower() == "y":
        if angkatan_siswa == "2021":
            jumlah_tagihan = 110000
        elif angkatan_siswa == "2022":
            jumlah_tagihan = 120000
        elif angkatan_siswa == "2023":
            jumlah_tagihan = 130000
        else:
            jumlah_tagihan = 140000
    elif metode_pilihan.lower() == "n":    
        try:
            jumlah_tagihan = float(input("Masukkan jumlah tagihan: "))
        except ValueError:
            print("Jumlah tagihan harus berupa angka.")
            return tentukan_tagihan(angkatan_siswa)
    else:
        print("Input tidak valid. Silakan masukkan 'y' atau 'n'.")
        return tentukan_tagihan(angkatan_siswa)
    print("Jumlah tagihan: ", format_rupiah(jumlah_tagihan))
    print_formatted_separator(len(pesan))
    return jumlah_tagihan
    

# FUNGSI FORMAT RUPIAH
# Mengubah angka menjadi format rupiah.
def format_rupiah(angka):
    try:
        # Cek apakah angka sudah dalam format numerik
        if not isinstance(angka, (int, float)):
            angka = float(angka) 

        # Format angka ke dalam format rupiah
        return f"Rp {angka:,.2f}".replace(',', '.').replace('.', ',', 1) 
    except ValueError:
        return "Input tidak valid, tidak dapat mengubah ke format angka."
    except Exception as e:
        return f"Terjadi kesalahan: {e}"


# FUNGSI UNTUK MENCETAK GARIS PEMBATAS
def print_formatted_separator(length=39):
    print("-" * length)