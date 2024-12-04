import openpyxl
import os
from datetime import datetime
from function import create_sheet_if_not_exists, print_formatted_separator, format_rupiah
from menu_tagihan import tampil_tagihan_belum_dibayar

FILE_NAME = 'data_spp.xlsx'  


# FUNGSI PEMBAYARAN SPP
# Menambahkan data pembayaran sekaligus mengubah status tagihan menjadi Sudah Dibayar
# Membayar tagihan berdasarkan kode tagihan, user menginput nominal uang.
def bayar_spp():
    # Memeriksa apakah file Excel ada
    if not os.path.exists(FILE_NAME):
        print("Data SPP belum ada.")
        return

    workbook = openpyxl.load_workbook(FILE_NAME)
    
    if 'Tagihan' not in workbook.sheetnames:
        print("Data Tagihan tidak ada.")
        return

    tagihan_sheet = workbook["Tagihan"]

    # Menampilkan data tagihan yang belum dibayar
    if not tampil_tagihan_belum_dibayar():
        return

    # Input data pembayaran
    kode_tagihan = input("Masukkan kode tagihan: ")

    # Input nominal uang
    try:
        nominal_uang = float(input("Masukkan nominal uang: "))
    except ValueError:
        print("Nominal uang harus berupa angka.")
        return
    
    tanggal_pembayaran = datetime.today().strftime('%Y-%m-%d')  

    tagihan_valid = False
    jumlah_tagihan = 0
    no_siswa = 0
    bulan_spp = ""
    status_tagihan = ""
    # Mencari tagihan berdasarkan kode tagihan
    for row in tagihan_sheet.iter_rows(min_row=2):
        if row[0].value == kode_tagihan:
            tagihan_valid = True
            no_siswa = row[1].value
            jumlah_tagihan = row[2].value  # Mengambil jumlah tagihan
            if row[4].value == "Sudah Dibayar":
                print("Tagihan sudah dibayar.")
                return
            break
        
    if not tagihan_valid:
        print("Tagihan tidak ditemukan.")
        return

    # Menghitung kembalian jika ada
    kembalian = nominal_uang - jumlah_tagihan
    if kembalian < 0:
        print("Nominal uang tidak cukup untuk membayar tagihan.")
        return

    # Menambahkan data pembayaran
    pembayaran_sheet = create_sheet_if_not_exists(workbook, 'Pembayaran', ['Kode Tagihan', 'Nomor Siswa', 'Tanggal Pembayaran', 'Jumlah tagihan', 'Uang Diterima', 'Kembalian'])
    pembayaran_sheet.append([kode_tagihan, no_siswa, tanggal_pembayaran, jumlah_tagihan, nominal_uang, kembalian])

    # Mengubah status tagihan menjadi "Sudah Dibayar"
    for row in tagihan_sheet.iter_rows(min_row=2):
        if row[0].value == kode_tagihan:
            row[4].value = "Sudah Dibayar"  # Mengubah status tagihan
            bulan_spp = row[3].value
            status_tagihan = row[4].value
            break

    workbook.save(FILE_NAME)
    print("Pembayaran berhasil...")
    print_formatted_separator()
    print("\nKode tagihan: ", kode_tagihan)
    print("Nomor siswa: ", no_siswa)
    print("Jumlah tagihan: ", format_rupiah(jumlah_tagihan))
    print("Bulan SPP: ", bulan_spp)
    print("Status tagihan: ", status_tagihan)
    print("Tanggal Pembayaran: ", tanggal_pembayaran)
    print("Kembalian Anda: ", kembalian)
    print_formatted_separator()



# FUNGSI MENAMPILKAN DATA PEMBAYARAN
# Menampilkan daftar data pembayaran
def tampil_data_pembayaran():
    if not os.path.exists(FILE_NAME):
        print("File data SPP tidak ditemukan.")
        return

    workbook = openpyxl.load_workbook(FILE_NAME)  
    pembayaran_sheet = create_sheet_if_not_exists(workbook, 'Pembayaran')

    if pembayaran_sheet.max_row == 1:
        print("Belum ada data pembayaran.")
        return

    print("\nDaftar Pembayaran:")
    for row in pembayaran_sheet.iter_rows(min_row=2, values_only=True):
        print(row)
    workbook.close()



# FUNGSI UNTUK MENAMPILKAN MENU PEMBAYARAN
# Menampilkan menu untuk mengelola pembayaran.
def menu_pembayaran():
    while True:
        print("\nMenu Pembayaran SPP:")
        print("1. Bayar SPP")
        print("2. Tampilkan daftar pembayaran")
        print("3. Kembali ke Menu Utama")
        sub_pilihan = input("Pilih menu: ")

        if sub_pilihan == '1':
            bayar_spp()
        elif sub_pilihan == '2':
            tampil_data_pembayaran()
        elif sub_pilihan == '3':
            break 
        else:
            print("Pilihan tidak valid. Silakan coba lagi.")
