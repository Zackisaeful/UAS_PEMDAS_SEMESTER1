import openpyxl
import os
from datetime import datetime  
from function import create_sheet_if_not_exists, generate_kode_tagihan, pilih_bulan, tentukan_tagihan, format_rupiah, print_formatted_separator
from menu_siswa import tampil_siswa

FILE_NAME = 'data_spp.xlsx'


# FUNGSI MENAMBAH DATA TAGIHAN BARU
# Menambahkan tagihan baru ke dalam daftar tagihan.
def tambah_tagihan():
    # Membuka atau membuat workbook
    if not os.path.exists(FILE_NAME):
        workbook = openpyxl.Workbook() 
        workbook.save(FILE_NAME) 
    else:
        workbook = openpyxl.load_workbook(FILE_NAME)

    # Membuat sheet tagihan jika tidak ada
    sheet = create_sheet_if_not_exists(workbook, 'Tagihan', ['Kode Tagihan', 'No Siswa', 'Jumlah Tagihan', 'Bulan SPP', 'Status Tagihan', 'Tanggal'])

    # Tampilkan data siswa untuk mempermudah user menentukan penagihan
    tampil_siswa()

    # Input data tagihan
    # Input nomor siwa
    no_siswa = input("\nMasukkan nomor siswa: ").strip()
    print_formatted_separator()
    if not no_siswa:
        print("Nomor siswa tidak boleh kosong.")
        return

    # Validasi keberadaan sheet Siswa
    if 'Siswa' not in workbook.sheetnames:
        print("Data siswa tidak ada, tidak dapat menambah tagihan.")
        return

    # Cek apakah siswa yang dimaksud ada 
    siswa_sheet = workbook["Siswa"]
    angkatan_siswa = None
    siswa_valid = False

    for row in siswa_sheet.iter_rows(min_row=2):
        if str(row[0].value) == no_siswa:
            siswa_valid = True
            angkatan_siswa = row[2].value
            break    

    if not siswa_valid:
        print("Siswa tidak terdaftar, mohon untuk memasukkan nomor yang benar.")
        return

    # Generate kode tagihan 
    kode_tagihan = generate_kode_tagihan(workbook)

    # Input bulan SPP
    bulan_spp = pilih_bulan()

    # Input jumlah tagihan
    jumlah_tagihan = tentukan_tagihan(angkatan_siswa) 
 
    
    # Status dan tanggal penagihan
    status_tagihan = "Belum Dibayar"
    tanggal = datetime.today().strftime('%Y-%m-%d') 

    # Menambahkan data tagihan ke sheet
    sheet.append([kode_tagihan, no_siswa, jumlah_tagihan, bulan_spp, status_tagihan, tanggal])
    workbook.save(FILE_NAME)
    workbook.close()  
    print("\nTagihan berhasil ditambahkan...")
    print_formatted_separator()
    print("\nKode tagihan: ", kode_tagihan)
    print("Nomor siswa: ", no_siswa)
    print("Jumlah tagihan: ", format_rupiah(jumlah_tagihan))
    print("Bulan SPP: ", bulan_spp)
    print("Status tagihan: ", status_tagihan)
    print("Tanggal: ", tanggal)
    print_formatted_separator()



# FUNGSI MENAMPILKAN SELURUH DATA TAGIHAN
# Menampilkan seluruh daftar tagihan
def tampil_tagihan():
    if not os.path.exists(FILE_NAME):
        print("File data SPP tidak ditemukan.")
        return

    workbook = openpyxl.load_workbook(FILE_NAME)
    if 'Tagihan' not in workbook.sheetnames:
        print("Sheet Tagihan belum ada.")
        return

    sheet = workbook['Tagihan']
    if sheet.max_row == 1:
        print("Belum ada data tagihan.")
        return

    print("\nDaftar Tagihan:")
    for row in sheet.iter_rows(min_row=2, values_only=True):
        print(row)
    workbook.close()



# FUNGSI MENAMPILKAN DATA TAGIHAN YANG BELUM DIBAYAR
# Menampikan daftar tagihan yang statusnya belum dibayar
def tampil_tagihan_belum_dibayar():
    if not os.path.exists(FILE_NAME):
        print("File data SPP tidak ditemukan.")
        return False

    workbook = openpyxl.load_workbook(FILE_NAME)
    if 'Tagihan' not in workbook.sheetnames:
        print("Sheet Tagihan belum ada.")
        return False

    sheet = workbook['Tagihan']
    if sheet.max_row == 1:
        print("\nBelum ada data tagihan.")
        return False

    print("\nDaftar Tagihan yang Belum Dibayar:")
    tagihan_belum_dibayar_found = False
    for row in sheet.iter_rows(min_row=2, values_only=True):
        if row[4] == "Belum Dibayar":
            print(row)
            tagihan_belum_dibayar_found = True

    if not tagihan_belum_dibayar_found:
        print("\nTidak ada tagihan yang belum dibayar!")
    
    workbook.close()
    return tagihan_belum_dibayar_found




# FUNGSI MENAMPILKAN DATA TAGIHAN YANG SUDAH DIBAYAR
# Menampikan daftar tagihan yang statusnya sudah dibayar
def tampil_tagihan_sudah_dibayar():
    if not os.path.exists(FILE_NAME):
        print("File data SPP tidak ditemukan.")
        return

    workbook = openpyxl.load_workbook(FILE_NAME)
    if 'Tagihan' not in workbook.sheetnames:
        print("Sheet Tagihan belum ada.")
        return

    sheet = workbook['Tagihan']
    if sheet.max_row == 1:
        print("Belum ada data tagihan.")
        return

    print("\nDaftar Tagihan yang Sudah Dibayar:")
    tagihan_sudah_dibayar_found = False  # Jika false berarti tidak ada data tagihan yang Sudah dibayar
    for row in sheet.iter_rows(min_row=2, values_only=True):
        if row[4] == "Sudah Dibayar":  # Memeriksa status tagihan
            print(row)
            tagihan_sudah_dibayar_found = True

    if not tagihan_sudah_dibayar_found:
        print("\nTidak ada tagihan yang sudah dibayar!")
        return 
    
    workbook.close()
    return tagihan_sudah_dibayar_found



# FUNGSI MENGEDIT DATA TAGIHAN
# Mengedit data tagihan yang sudah terdaftar berdasarkan kode tagihan.
def edit_tagihan():
    # Tampilkan data siswa untuk mempermudah user menentukan penagihan
    tampil_tagihan()

    if not os.path.exists(FILE_NAME):
        print("File data SPP tidak ditemukan.")
        return

    workbook = openpyxl.load_workbook(FILE_NAME)
    
    # Pastikan sheet 'Tagihan' ada
    if 'Tagihan' not in workbook.sheetnames:
        print("Sheet Tagihan belum ada.")
        workbook.close()
        return
    
    # Pastikan sheet 'Siswa' ada
    if 'Siswa' not in workbook.sheetnames:
        print("Sheet Siswa belum ada.")
        workbook.close()
        return

    sheet_tagihan = workbook['Tagihan']
    sheet_siswa = workbook['Siswa']

    kode_tagihan = input("\nMasukkan kode tagihan yang ingin diedit: ").strip()

    # Cari tagihan berdasarkan kode
    tagihan_row = None
    for row in sheet_tagihan.iter_rows(min_row=2, values_only=False):
        if str(row[0].value) == kode_tagihan:
            tagihan_row = row
            break

    if not tagihan_row:
        print("Tagihan tidak ditemukan.")
        workbook.close()
        return

    # Input nomor siswa baru
    no_siswa_baru = input("\nMasukkan nomor siswa baru: ").strip()
    if not no_siswa_baru:
        print("Nomor siswa tidak boleh kosong.")
        workbook.close()
        return

    # Validasi apakah nomor siswa ada di sheet 'Siswa'
    siswa_valid = False
    angkatan_siswa = None
    for row in sheet_siswa.iter_rows(min_row=2, values_only=True):
        if str(row[0]) == no_siswa_baru:
            siswa_valid = True
            angkatan_siswa = row[2]  
            break

    if not siswa_valid:
        print("Siswa tidak terdaftar. Masukkan nomor siswa yang benar.")
        workbook.close()
        return

    try:
        jumlah_tagihan = tentukan_tagihan(angkatan_siswa)
    except ValueError:
        print("Jumlah tagihan harus berupa angka.")
        workbook.close()
        return

    bulan_spp = pilih_bulan()
    tanggal = datetime.today().strftime('%Y-%m-%d') 

    tagihan_row[1].value = no_siswa_baru  
    tagihan_row[2].value = jumlah_tagihan 
    tagihan_row[3].value = bulan_spp  
    tagihan_row[5].value = tanggal

    print("Tagihan berhasil diedit.")
    workbook.save(FILE_NAME)
    workbook.close()



# FUNGSI MENGHAPUS DATA TAGIHAN
# Menghapus data tagihan yang sudah terdaftar berdasarkan kode tagihan
def hapus_tagihan():
    kode_tagihan = input("Masukkan kode tagihan yang ingin dihapus: ")

    if not os.path.exists(FILE_NAME):
        print("File data SPP tidak ditemukan.")
        return

    workbook = openpyxl.load_workbook(FILE_NAME)
    if 'Tagihan' not in workbook.sheetnames:
        print("Sheet Tagihan belum ada.")
        return

    sheet = workbook['Tagihan']
    rows_to_delete = []
    for row_index, row in enumerate(sheet.iter_rows(min_row=2), start=2):
        if str(row[0].value) == kode_tagihan:
            rows_to_delete.append(row_index)

    if rows_to_delete:
        for row_index in sorted(rows_to_delete, reverse=True):
            sheet.delete_rows(row_index)  
        print("Tagihan berhasil dihapus.")
    else:
        print("Tagihan tidak ditemukan.")

    workbook.save(FILE_NAME)



# FUNGSI MENAMPILKAN MENU TAGIHAN
# Menampilkan menu untuk mengelola data tagihan.
def menu_tagihan():
    while True:
        print("\nMenu Tagihan:")
        print("1. Tambah Tagihan")
        print("2. Tampil seluruh Tagihan")
        print("3. Tampil seluruh Tagihan yang belum dibayar")
        print("4. Tampil seluruh Tagihan yang sudah dibayar")
        print("5. Edit Tagihan")
        print("6. Hapus Tagihan")
        print("7. Kembali ke Menu Utama")
        sub_pilihan = input("Pilih menu: ")

        if sub_pilihan == '1':
            tambah_tagihan()
        elif sub_pilihan == '2':
            tampil_tagihan()
        elif sub_pilihan == '3':
            tampil_tagihan_belum_dibayar()
        elif sub_pilihan == '4':
            tampil_tagihan_sudah_dibayar()
        elif sub_pilihan == '5':
            edit_tagihan()
        elif sub_pilihan == '6':
            hapus_tagihan()
        elif sub_pilihan == '7':
            break
        else:
            print("Pilihan tidak valid. Silakan coba lagi.")