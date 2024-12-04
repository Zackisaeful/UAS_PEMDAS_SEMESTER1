import openpyxl
import os
from function import create_sheet_if_not_exists, generate_no_siswa, tahun_angkatan, print_formatted_separator

FILE_NAME = 'data_spp.xlsx'

# FUNGSI MENAMBAH DATA SISWA 
# Menambahkan siswa baru ke dalam daftar siswa.
def tambah_siswa():   
    if not os.path.exists(FILE_NAME):
        workbook = openpyxl.Workbook()
    else:
        workbook = openpyxl.load_workbook(FILE_NAME)

    sheet = create_sheet_if_not_exists(workbook, 'Siswa', ['No Siswa', 'Nama Siswa', 'Tahun Angkatan', 'Kelas', 'No HP', 'Alamat'])

    # Menghasilkan nomor siswa baru
    no_siswa = generate_no_siswa(workbook)
   
    # Memeriksa apakah nomor siswa sudah ada
    for row in sheet.iter_rows(min_row=2):
        if row[0].value == no_siswa:
            print("Nomor siswa sudah ada. Silakan coba lagi.")
            return

    # Input data siswa
    # Input nama siswa
    # Method strip() menghapus  karakter whitespace
    nama_siswa = input("\nMasukkan nama siswa: ").strip()
    print_formatted_separator()
    if not nama_siswa:
        print("\nNama siswa tidak boleh kosong.")
        return

    # Input tahun angkatan
    angkatan_siswa = tahun_angkatan().strip()

    # Input kelas siswa
    kelas = input("\nMasukan kelas: ").strip()
    print_formatted_separator()
    if not kelas:
        print("\nKelas tidak boleh kosong.")
        return    

    # Input nomor hp siswa
    no_hp = input("\nMasukkan nomor HP siswa: ").strip()
    print_formatted_separator()
    if not no_hp:
        print("\nNomor hp tidak boleh kosong.")
        return    

    # Input alamat siswa
    alamat = input("\nMasukkan alamat siswa: ").strip()
    print_formatted_separator()
    if not alamat:
        print("\nAlamat tidak boleh kosong.")
        return

    sheet.append([no_siswa, nama_siswa, angkatan_siswa, kelas, no_hp, alamat])
    workbook.save(FILE_NAME)
    print("\nSiswa berhasil ditambahkan...")
    print_formatted_separator()
    print("\nNomor Siswa: ", no_siswa)
    print("Nama: ", nama_siswa)
    print("Kelas: ", kelas )
    print("Angkatan: ", angkatan_siswa )
    print("Nomor Hp: ", no_hp )
    print("Alamat: ", alamat )
    print_formatted_separator()



# FUNGSI MENAMPILKAN SELURUH DATA SISWA
# Menampilkan daftar siswa yang terdaftar.
def tampil_siswa():
    if not os.path.exists(FILE_NAME):
        print("File data SPP tidak ditemukan.")
        return

    workbook = openpyxl.load_workbook(FILE_NAME)
    if 'Siswa' not in workbook.sheetnames:
        print("Sheet Siswa belum ada.")
        return

    sheet = workbook['Siswa']
    if sheet.max_row == 1:
        print("Belum ada data siswa.")
        return

    print("\nDaftar Siswa:")
    for row in sheet.iter_rows(min_row=2, values_only=True):
        print(row)
    workbook.close()



# FUNGSI EDIT SISWA
#  Mengedit data siswa yang sudah terdaftar berdasarkan nomor siswa.
def edit_siswa():
    no_siswa = input("Masukkan nomor siswa yang ingin diedit: ")

    if not os.path.exists(FILE_NAME):
        print("File data SPP tidak ditemukan.")
        return

    workbook = openpyxl.load_workbook(FILE_NAME)
    if 'Siswa' not in workbook.sheetnames:
        print("Sheet Siswa belum ada.")
        return

    sheet = workbook['Siswa']
    for row in sheet.iter_rows(min_row=2):
        if  str(row[0].value) == no_siswa:
            nama_siswa = input("Masukkan nama siswa: ")
            angkatan_siswa = tahun_angkatan()
            kelas = input("Masukan kelas: ")
            no_hp = input("Masukkan nomor HP siswa: ")
            alamat = input("Masukkan alamat siswa: ")
            row[1].value = nama_siswa
            row[2].value = angkatan_siswa
            row[3].value = kelas
            row[4].value = no_hp
            row[5].value = alamat
            print("Siswa berhasil diedit.")
            workbook.save(FILE_NAME)
            return

    print("Siswa tidak ditemukan.")
    workbook.save(FILE_NAME)



# FUNGSI HAPUS DATA SISWA
# Menghapus siswa dari daftar berdasarkan nomor siswa
def hapus_siswa():
    no_siswa = input("Masukkan nomor siswa yang ingin dihapus: ")

    if not os.path.exists(FILE_NAME):
        print("File data SPP tidak ditemukan.")
        return

    workbook = openpyxl.load_workbook(FILE_NAME)
    if 'Siswa' not in workbook.sheetnames:
        print("Sheet Siswa belum ada.")
        return

    sheet = workbook['Siswa']
    rows_to_delete = []
    for row_index, row in enumerate(sheet.iter_rows(min_row=2), start=2):
        if  str(row[0].value) == no_siswa:

            rows_to_delete.append(row_index)

    if rows_to_delete:
        for row_index in sorted(rows_to_delete, reverse=True):
            sheet.delete_rows(row_index)  
        print("Siswa berhasil dihapus.")
    else:
        print("Siswa tidak ditemukan.")

    workbook.save(FILE_NAME)



# FUNGSI MENU SISWA
# Menampilkan menu untuk mengelola data siswa.
def menu_siswa():
    while True:
        print("\nMenu Siswa:")
        print("1. Tambah Siswa")
        print("2. Tampil Siswa")
        print("3. Edit Siswa")
        print("4. Hapus Siswa")
        print("5. Kembali ke Menu Utama")
        sub_pilihan = input("Pilih menu: ")

        if sub_pilihan == '1':
            tambah_siswa()
        elif sub_pilihan == '2':
            tampil_siswa()
        elif sub_pilihan == '3':
            edit_siswa()
        elif sub_pilihan == '4':
            hapus_siswa()
        elif sub_pilihan == '5':
            break
